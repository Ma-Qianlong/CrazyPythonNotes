#!/usr/bin/env python

# -*- *************** -*-
# @File  : xlsx_template.py
# @Description : 
# @Author: mql
# @Time  : 2021/1/29 15:45
# -*- *************** -*-


# 2.操作xlsx格式的表格文件：
#
# 读取/写入：openpyxl

import openpyxl
from xlutils.copy import copy


def write_excel_xlsx(path, sheet_name, value):
    index = len(value)
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = sheet_name
    for i in range(0, index):
        for j in range(0, len(value[i])):
            sheet.cell(row=i + 1, column=j + 1, value=str(value[i][j]))
    workbook.save(path)
    print("xlsx格式表格写入数据成功！")

def write_excel_xlsx_append(path, sheet_name, value):
    workbook = openpyxl.load_workbook(path)
    # sheet = wb.get_sheet_by_name(sheet_name)这种方式已经弃用，不建议使用
    sheet = workbook[sheet_name]
    max_row = sheet.max_row
    for i in range(0, len(value)):
        for j in range(0, len(value[i])):
            sheet.cell(row=max_row + i + 1, column=j + 1, value=str(value[i][j]))
    workbook.save(path)
    print("xlsx格式表格【追加】写入数据成功!")

def read_excel_xlsx(path, sheet_name):
    workbook = openpyxl.load_workbook(path)
    # sheet = wb.get_sheet_by_name(sheet_name)这种方式已经弃用，不建议使用
    sheet = workbook[sheet_name]
    max_row = sheet.max_row
    for row in sheet.rows:
        for cell in row:
            print(cell.value, "\t", end="")
        print()

    # # 可以直接追加写数据
    # sheet.cell(row=max_row + 1, column=1, value=str("test"))
    # workbook.save(path)

if __name__ == '__main__':
    book_name_xlsx = 'xlsx格式测试工作簿.xlsx'

    sheet_name_xlsx = 'xlsx格式测试表'

    value3 = [["姓名", "性别", "年龄", "城市", "职业"],
              ["111", "女", "66", "石家庄", "运维工程师"],
              ["222", "男", "55", "南京", "饭店老板"],
              ["333", "女", "27", "苏州", "保安"], ]

    write_excel_xlsx(book_name_xlsx, sheet_name_xlsx, value3)
    read_excel_xlsx(book_name_xlsx, sheet_name_xlsx)
