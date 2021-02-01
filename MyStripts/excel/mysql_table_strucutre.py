#!/usr/bin/env python

# -*- *************** -*-
# @File  : mysql_table_strucutre.py
# @Description : 
# @Author: mql
# @Time  : 2021/1/29 16:48
# -*- *************** -*-


import pymysql
import openpyxl


def selectTableNameAndComments(conn, database):
    '''
    根据库名获取所有表名称和表说明
    :return:
    '''
    cursor = conn.cursor()
    try:
        # sql = "SELECT TABLE_NAME, TABLE_COMMENT FROM information_schema.`TABLES` WHERE TABLE_SCHEMA = '" + database + "' and TABLE_COMMENT <> ''"
        sql = "SELECT TABLE_NAME, TABLE_COMMENT FROM information_schema.`TABLES` WHERE TABLE_SCHEMA = '" + database + "'"

        cursor.execute(sql)

        # logger.debug('---------------------------------------------')
        # a = 0;
        # for dd in cursor.description:
        #     logger.debug(a)
        #     a += 1
        #     logger.debug(dd)
        # logger.debug('---------------------------------------------')

        print('【selectTableNameAndComments】共查询到：%d条数据。' % cursor.rowcount)

        # 查询所有数据，返回结果默认以元组形式，所以可以进行迭代处理
        # for i in cursor.fetchall():
        #     print(i)

        # 获取第一行数据
        # result_1 = cursor.fetchone()
        # logger.debug(result_1)

        # 获取前n行数据
        # result_3 = cursor.fetchmany(3)
        # logger.debug(result_3)

        return cursor.fetchall()
    except Exception as e:
        print("【selectTableNameAndComments】 error, args: %s" % e.args)
    finally:
        cursor.close()


def selectTableCloumnInfo(conn, database, tablename):
    """
    获取指定表的列信息
    :param conn:
    :param database:
    :param tablename:
    :return:
    """

    cursor = conn.cursor()
    try:
        sql = """SELECT
            COLUMN_NAME AS '列名',
            ORDINAL_POSITION AS '列的排列顺序',
            COLUMN_DEFAULT AS '默认值',
            IS_NULLABLE AS '是否为空',
            DATA_TYPE AS '数据类型',
            CHARACTER_MAXIMUM_LENGTH AS '字符最大长度',
            NUMERIC_PRECISION AS '数值精度(最大位数)',
            NUMERIC_SCALE AS '小数精度',
            COLUMN_TYPE AS 列类型,
            COLUMN_KEY 'KEY',
            EXTRA AS '额外说明',
            COLUMN_COMMENT AS '注释'
        FROM
            information_schema.`COLUMNS`
        WHERE
            TABLE_SCHEMA = '""" + database + "' and TABLE_NAME = '" + tablename + "'" + """
        ORDER BY
            ORDINAL_POSITION;"""

        cursor.execute(sql)

        # logger.debug('---------------------------------------------')
        # a = 0;
        # for dd in cursor.description:
        #     logger.debug(a)
        #     a += 1
        #     logger.debug(dd)
        # logger.debug('---------------------------------------------')

        print('【selectTableCloumnInfo】共查询到：%d条数据。' % cursor.rowcount)

        # 查询所有数据，返回结果默认以元组形式，所以可以进行迭代处理
        # for i in cursor.fetchall():
        #     print(i)

        # 获取第一行数据
        # result_1 = cursor.fetchone()
        # logger.debug(result_1)

        # 获取前n行数据
        # result_3 = cursor.fetchmany(3)
        # logger.debug(result_3)

        return cursor.fetchall()
    except Exception as e:
        print("【selectTableCloumnInfo】 error, args: %s" % e.args)
    finally:
        cursor.close()


def write_excel_xlsx(path, sheet_name, value):
    index = len(value)
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = sheet_name
    for i in range(0, index):
        for j in range(0, len(value[i])):
            sheet.cell(row=i + 1, column=j + 1, value=str(value[i][j]))
    workbook.save(path)
    print("xlsx格式表格写入数据成功！")


def write_excel_xlsx_append(path, sheet_name, value):
    workbook = openpyxl.load_workbook(path)
    # sheet = wb.get_sheet_by_name(sheet_name)这种方式已经弃用，不建议使用
    sheet = workbook[sheet_name]
    max_row = sheet.max_row
    for i in range(0, len(value)):
        for j in range(0, len(value[i])):
            sheet.cell(row=max_row + i + 1, column=j + 1, value=str(value[i][j]))
    workbook.save(path)
    print("xlsx格式表格【追加】写入数据成功!")


conn = pymysql.connect(host='127.0.0.1', port=3136, user='root', password='root', database='db-psbc-credit-pay')


def mysqlTableStructureToExcel(database):
    tabname_comment = selectTableNameAndComments(conn, database)
    book_name_xlsx = "MqlTableInfo.xlsx"
    sheet_name_xls = "table info"
    write_excel_xlsx(book_name_xlsx, sheet_name_xls, [])

    for tc in tabname_comment:
        print(tc)
        excelVal = [[], ['中文名称', tc[1]], ['英文名称', tc[0]]]
        excelVal.append(['字段名称', '字段说明', '字段类型', '是否为空', 'KEY', '备注'])
        ci = selectTableCloumnInfo(conn, database, tc[0])
        for i in ci:
            excelVal.append([i[0], i[11], i[8], i[3], i[9]])

        write_excel_xlsx_append(book_name_xlsx, sheet_name_xls, excelVal)


if __name__ == '__main__':
    mysqlTableStructureToExcel('db-psbc-credit-pay')

    # selectTableNameAndComments(conn, 'db-psbc-credit-pay')
    # selectTableCloumnInfo(conn, 'db-psbc-credit-pay', 'sys_menu')
    # if conn:
    #     conn.close()

    # book_name_xls = 'xlsx格式测试工作簿1.xlsx'
    #
    # sheet_name_xls = 'xlsx格式测试表'
    #
    # value_title = [["姓名", "性别", "年龄", "城市", "职业"], ]
    #
    # value1 = [["张三", "男", "19", "杭州", "研发工程师"],
    #           ["李四", "男", "22", "北京", "医生"],
    #           ["王五", "女", "33", "珠海", "出租车司机"], ]
    #
    # value2 = [["Tom", "男", "21", "西安", "测试工程师"],
    #           ["Jones", "女", "34", "上海", "产品经理"],
    #           ["Cat", "女", "56", "上海", "教师"], ]
    #
    # write_excel_xlsx(book_name_xls, sheet_name_xls, value_title)
    # write_excel_xlsx_append(book_name_xls, sheet_name_xls, value1)
    # write_excel_xlsx_append(book_name_xls, sheet_name_xls, value2)
